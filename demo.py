import pandas as pd
import xlwings as xw
#%%
df = pd.read_csv('files/100_CC_Records.csv', encoding='ISO-8859-1')
grouped_bank_df = df.groupby(['Issuing Bank'])
excelbook = xw.Book()
for i, (name, group_df) in enumerate(grouped_bank_df):
    print(name)
    if i == 0:
        excelbook.sheets[i].range('A1').value = group_df 
    else:
        excelbook.sheets.add().range('A1').value = group_df

excelbook.save('files/s.xlsx')

#%%
import pandas as pd
import xlwings as xw
import shutil

#Merge new data
sales_df = pd.read_excel('files/100_Sales_Records.xlsx', 
                         sheet_name='100_Sales_Records', 
                         encoding='ISO-8859-1')
new_sales_df = pd.read_csv('files/1000_Sales_Records.csv', encoding='ISO-8859-1')
merged_df = pd.concat([sales_df, new_sales_df])

shutil.copy2('files/100_Sales_Records.xlsx', 'files/New_Sales_Records.xlsx')
# sales_excel = xw.Book('files/100_Sales_Records.xlsx')
try:
    app_excel = xw.App(visible=False, add_book=False)
    sales_excel = app_excel.books.open('files/New_Sales_Records.xlsx')
    app_excel.visible = False
    sales_sht = sales_excel.sheets['100_Sales_Records']
    sales_sht.range('A1').options(index=False).value = merged_df

# new_sales_excel = xw.Book('files/New_Sales_Records.xlsx')
# print(len(sales_excel.sheets))
# for i in range(len(sales_excel.sheets)):
#     sales_excel.sheets[i].api.Copy(Before=new_sales_excel.sheets[i].api)

# new_sales_excel.save('files/New_Sales_Records.xlsx')

    sales_excel.api.RefreshAll()
    sales_excel.save()
# sales_excel.save('files/new.xlsx')

    # sales_excel.sheets['PVT'].select()
    # sales_excel.api.ActiveSheet.PivotTables('PVT').PivotCache().refresh()
# sales_excel.close()
# app_excel.quit()
# rows = sales_sht.api.UsedRange.Rows.count
# cols = sales_sht.api.UsedRange.Columns.count
except Exception as e:
    print('error msg:{}'.format(e))





#%%
import pandas as pd
import xlwings as xw
import shutil
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

#Merge new data
sales_df = pd.read_excel('files/100_Sales_Records.xlsx', 
                         sheet_name='100_Sales_Records', 
                         encoding='ISO-8859-1')
new_sales_df = pd.read_csv('files/1000_Sales_Records.csv', encoding='ISO-8859-1')
merged_df = pd.concat([sales_df, new_sales_df])

shutil.copy2('files/100_Sales_Records.xlsx', 'files/New_Sales_Records.xlsx')
wb = load_workbook(filename = 'files/New_Sales_Records.xlsx')
ws = wb['100_Sales_Records']
rows = dataframe_to_rows(merged_df, index=False)

for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
         ws.cell(row=r_idx, column=c_idx, value=value)

ws = wb['PVT']
pivot = ws._pivots[0]
pivot.cache.refreshOnload = True
wb.save('files/New_Sales_Records.xlsx')


#%%
import pandas as pd
import xlwings as xw
import shutil
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

#Merge new data
sales_df = pd.read_excel('files/100_Sales_Records_Formula.xlsx', 
                         sheet_name='100_Sales_Records', 
                         encoding='ISO-8859-1')
new_sales_df = pd.read_csv('files/1000_Sales_Records.csv', encoding='ISO-8859-1')
new_sales_df['New Order'] = new_sales_df['Order Priority'].map(str) + new_sales_df['Order ID'].map(str)
merged_df = pd.concat([sales_df, new_sales_df])

shutil.copy2('files/100_Sales_Records_Formula.xlsx', 'files/New_Sales_Records_Formula.xlsx')
wb = load_workbook(filename = 'files/New_Sales_Records_Formula.xlsx')
ws = wb['100_Sales_Records']
rows = dataframe_to_rows(merged_df, index=False)

for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
         ws.cell(row=r_idx, column=c_idx, value=value)

ws = wb['PVT']
pivot = ws._pivots[0]
pivot.cache.refreshOnload = True
wb.save('files/New_Sales_Records_Formula.xlsx')

#%%

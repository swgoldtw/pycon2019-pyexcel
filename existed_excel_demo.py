#%% Demo: Update cross worksheet formula cells
import pandas as pd
import shutil
import math
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.cell import get_column_letter
from openpyxl.formula.translate import Translator
from itertools import islice
from excel_helper import convert_ws_df, refresh_pv

print('----load excel file----')
wb = load_workbook(filename = 'files/excel_from_boss.xlsx')

# merge sales_record data
print('----merge sales_record data----')
sales_df = convert_ws_df(wb['sales_records'], True)
new_sales_df = pd.read_csv('files/1000_Sales_Records.csv', encoding='ISO-8859-1')
merged_df = pd.concat([sales_df, new_sales_df], sort=False)

# update sales_records worksheet
print('----apply formula for appended data----')
sales_ws = wb['sales_records']
rows = dataframe_to_rows(merged_df, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        if isinstance(value, float) and math.isnan(value):
            source_cell_idx = str(get_column_letter(c_idx)) + str(r_idx-1)
            target_cell_idx = str(get_column_letter(c_idx)) + str(r_idx)
            trans = Translator(sales_ws.cell(row=r_idx-1,column=c_idx).value, origin=source_cell_idx)
            value = trans.translate_formula(target_cell_idx)
        sales_ws.cell(row=r_idx, column=c_idx, value=value)

# merge excel chart data
print('----merge chart data----')
chart_df = convert_ws_df(wb['chart_sample'], True)
new_chart_df = pd.read_csv('files/Chart_Sales_Records.csv', encoding='ISO-8859-1')
merged_df = pd.concat([chart_df, new_chart_df], sort=False)

# update chart worksheet
print('----refresh worksheet chart----')
chart_ws = wb['chart_sample']
rows = dataframe_to_rows(merged_df, index=False)
for r_idx,row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        chart_ws.cell(row=r_idx, column=c_idx, value=value)

# refresh pivot table
refresh_pv(wb['PVT'])

# save updates
wb.save('files/excel_to_boss.xlsx')


#%%
